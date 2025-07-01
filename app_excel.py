import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import datetime
from PIL import Image
from streamlit_drawable_canvas import st_canvas

# Definir campos obligatorios por tipo de servicio
CAMPOS_OBLIGATORIOS_POR_TIPO = {
    "Mantención": [
        "cotizacion", "orden_compra", "tecnico", "ppm", "corriente", "voltaje",
        "presion_entrada", "presion_salida", "observaciones", "lista_piezas",
        "fotos_antes", "fotos_despues", "recibe_nombre", "recibe_email",
        "recibe_cargo", "recibe_telefono"
    ],
    "Instalación": [
        "cotizacion", "orden_compra", "factura", "tecnico", "ppm", "corriente",
        "voltaje", "presion_entrada", "presion_salida", "observaciones",
        "capacitacion_si", "lista_piezas", "boquillas_instaladas", "fotos_antes",
        "fotos_despues", "recibe_nombre", "recibe_email", "recibe_cargo", "recibe_telefono"
    ],
    "Urgencias": [
        "tecnico", "observaciones", "recibe_nombre", "recibe_cargo",
        "recibe_email", "recibe_telefono"
    ],
    "Post Ventas": [
        "tecnico", "ppm", "corriente", "voltaje", "presion_entrada",
        "presion_salida", "observaciones", "recibe_nombre", "recibe_email",
        "recibe_cargo", "recibe_telefono"
    ],
    "Venta Repuestos": [
        "tecnico", "observaciones", "recibe_nombre", "recibe_email",
        "recibe_cargo", "recibe_telefono"
    ]
}


TEMPLATE_PATH = "IT Tipo (002).xlsx"

# Configuración de la página
st.set_page_config(page_title="\n\nGenerador de Informe Técnico", layout="wide")

def cargar_logo_por_tema():
    try:
        # Detectar tema solo una vez
        if 'modo_tema' not in st.session_state:
            modo_detectado = st.get_option("theme.base")
            st.session_state.modo_tema = modo_detectado if modo_detectado else "dark"  # Preferir dark si es desconocido

        modo = st.session_state.modo_tema

        if modo == "dark":
            return Image.open("Logo_blanco.png")
        else:
            return Image.open("Logo_negro.png")
    except Exception:
        return None

# Agregar el logo y el título en columnas
col_title, col_logo  = st.columns([4,3])

with col_logo:
    logo = cargar_logo_por_tema()
    if logo:
        st.image(logo, width=500)
    else:
        st.warning("⚠️ No se pudo cargar el logo.")


with col_title:
    st.markdown("<br><br>", unsafe_allow_html=True)  # Agrega dos líneas de espacio
    st.title("\n\nGenerador de Informe Técnico")

st.markdown("---")


# === Leer opciones desde hoja 'Lista Servicio' (A2:A5, técnicos en G2:G10) ===
wb_temp = load_workbook(TEMPLATE_PATH, data_only=True)
ws_lista = wb_temp["Lista Servicio"]
tipo_opciones = [ws_lista[f"A{row}"].value for row in range(2, 7)]
tecnicos_opciones = [ws_lista[f"G{row}"].value for row in range(2, 11) if ws_lista[f"G{row}"].value]

# === Leer datos de clientes desde hoja "Clientes" ===
ws_clientes = wb_temp["Clientes"]
clientes_data = {}
for row in ws_clientes.iter_rows(min_row=2, values_only=True):
    if row[0]:
        clientes_data[str(row[0])] = {
            "razon_social": row[1],
            "rut": row[2],
            "ubicacion": row[3],
            "modelo": row[6],
        }
serie_opciones = list(clientes_data.keys())
wb_temp.close()

# === 🆔 Identificación del Informe ===
st.markdown("### 🆔 Identificación del Informe")
col_id1, col_id2 = st.columns(2)
with col_id1:
    it_num = st.text_input("N° IT")
    cotizacion = st.text_input("Cotización")
with col_id2:
    orden_compra = st.text_input("Orden de Compra")
    factura = st.text_input("N° Factura")

# === Técnico responsable ===
tecnico = st.selectbox("👨‍🔧 Técnico Responsable", tecnicos_opciones)

# === 1. Entradas generales ===
col1, col2 = st.columns(2)

with col1:
    tipo = st.selectbox("Tipo", tipo_opciones)
    fecha = st.date_input("Fecha", value=datetime.date.today())
    serierack = st.text_input("N° Serie Rack")
    serie = st.selectbox("N° Serie Equipo", ["" + s for s in serie_opciones])
    modelo = razon_social = rut_cliente = ubicacion = ""
    if serie and serie in clientes_data:
        cliente_info = clientes_data[serie]
        modelo = cliente_info["modelo"]
        razon_social = cliente_info["razon_social"]
        rut_cliente = cliente_info["rut"]
        ubicacion = cliente_info["ubicacion"]
    st.text(f"Modelo: {modelo}")
    st.text(f"Razón Social: {razon_social}")
    st.text(f"RUT Cliente: {rut_cliente}")
    st.text(f"Ubicación Física: {ubicacion}")
    observaciones = st.text_area("Observaciones Generales")

    capacitacion_si = st.radio("¿Hubo capacitación?", ["No", "Sí"])
    if "cap_list" not in st.session_state:
        st.session_state.cap_list = []

    if capacitacion_si == "Sí":
        st.markdown("**Personas capacitadas**")
        if len(st.session_state.cap_list) < 3:
            with st.form("form_capacitados", clear_on_submit=False):
                nombre = st.text_input("Nombre")
                rut = st.text_input("Cargo")
                submit = st.form_submit_button("+ Agregar")
                if submit and nombre and rut:
                    st.session_state.cap_list.append((nombre, rut))
        else:
            st.info("✅ Se alcanzó el límite de 3 personas capacitadas.")

        for i, (n, r) in enumerate(st.session_state.cap_list):
            cols = st.columns([5, 1])
            cols[0].write(f"{i+1}. {n} ({r})")
            if cols[1].button("❌", key=f"del_{i}"):
                st.session_state.cap_list.pop(i)
                st.experimental_rerun()

with col2:
    ppm = st.text_input("PPM Agua")
    agua = st.text_input("Origen del agua")
    corriente = st.text_input("Corriente de Trabajo")
    voltaje = st.text_input("Voltaje de Trabajo")
    presion_entrada = st.text_input("Presión de Entrada")
    presion_salida = st.text_input("Presión de Salida")
    fotos_antes = st.file_uploader("📸 Imagen 1", type=["jpg", "jpeg", "png"], key="foto_antes")
    fotos_despues = st.file_uploader("📸 Imagen 2", type=["jpg", "jpeg", "png"], key="foto_despues")
    #fotos = st.text_area("Comentario Fotográfico")

# === Boquillas Instaladas ===
st.markdown("### 🔧 Boquillas")
boquillas_instaladas = st.number_input(
    "Cantidad de Boquillas Instaladas",
    min_value=0,
    max_value=500,  # Puedes ajustar este máximo según necesites
    value=0,
    step=1,
    help="Ingrese el número total de boquillas instaladas"
)


# === 2. Piezas utilizadas ===
st.markdown("### 🧩 Piezas utilizadas")

# Cargar lista de piezas desde el excel
wb_temp = load_workbook(TEMPLATE_PATH, data_only=True)
ws_piezas = wb_temp["Piezas y Partes"]
piezas_data = {}
for row in ws_piezas.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1]:
        piezas_data[row[1]] = {"codigo": row[0]}

# Inicializar la lista de piezas en el estado de la sesión si no existe
if 'lista_piezas' not in st.session_state:
    st.session_state.lista_piezas = []


def agregar_pieza():
    if st.session_state.nueva_pieza:
        nueva_pieza = {
            "codigo": piezas_data[st.session_state.nueva_pieza]["codigo"],
            "nombre": st.session_state.nueva_pieza,
            "cantidad": st.session_state.nueva_cantidad,
            "garantia": st.session_state.nueva_garantia
        }
        st.session_state.lista_piezas.append(nueva_pieza)


# Formulario para agregar nueva pieza
with st.expander("➕ Agregar Nueva Pieza", expanded=True):
    col1, col2, col3 = st.columns([2, 1, 1])

    with col1:
        nombre_pieza = st.selectbox(
            "Nombre de la Pieza",
            options=[""] + list(piezas_data.keys()),
            key="nueva_pieza",
            index=0
        )

    with col2:
        cantidad = st.number_input(
            "Cantidad",
            min_value=1,
            value=1,
            key="nueva_cantidad"
        )

    with col3:
        garantia = st.selectbox(
            "Garantía",
            options=["Sí", "No"],
            key="nueva_garantia"
        )

    # Botón para agregar la pieza
    if st.button("✅ Agregar", use_container_width=True, on_click=agregar_pieza):
        st.rerun()

# Mostrar tabla de piezas agregadas
if st.session_state.lista_piezas:
    st.markdown("#### 📋 Piezas Agregadas")

    # Convertir la lista de piezas a DataFrame para mostrar
    df_piezas = pd.DataFrame(st.session_state.lista_piezas)

    # Mostrar tabla con las piezas
    st.data_editor(
        df_piezas,
        column_config={
            "codigo": st.column_config.TextColumn(
                "Código",
                disabled=True,
            ),
            "nombre": st.column_config.TextColumn(
                "Nombre de la Pieza",
                disabled=True,
            ),
            "cantidad": st.column_config.NumberColumn(
                "Cantidad",
                min_value=1,
                required=True,
            ),
            "garantia": st.column_config.SelectboxColumn(
                "Garantía",
                options=["Sí", "No"],
                required=True,
            ),
        },
        hide_index=True,
        key="tabla_piezas"
    )

    # Botón para eliminar la última pieza
    col1, col2 = st.columns([4, 1])
    with col2:
        if st.button(" Eliminar ", use_container_width=True):
            st.session_state.lista_piezas.pop()
            st.rerun()

else:
    st.info("No hay piezas agregadas.")

# === 3. Recibe conforme ===
st.markdown("### 📥 Recibe Conforme")
col_rec1, col_rec2 = st.columns(2)
with col_rec1:
    recibe_nombre = st.text_input("Nombre Cliente")
    recibe_email = st.text_input("Email")
with col_rec2:
    recibe_cargo = st.text_input("Cargo")
    recibe_telefono = st.text_input("Teléfono")

# === 4. Firma digital ===
st.markdown("### ✍️ Firma del Cliente")
sig_canvas = st_canvas(
    fill_color="#000000",          # Color negro para el trazo
    stroke_width=3,                # Grosor del trazo
    stroke_color="#000000",        # Color negro para la firma
    background_color="#ffffff",     # Fondo blanco
    height=148,
    width=500,
    drawing_mode="freedraw",
    key="canvas_firma"
)


# Convertir técnico a mayúsculas antes de guardar
tecnico_upper = tecnico.upper()

# === Validación de campos obligatorios según tipo de servicio ===
if st.button("📥 Generar Informe Excel"):
    campos_requeridos = CAMPOS_OBLIGATORIOS_POR_TIPO[tipo]
    campos_faltantes = []

    # Verificar cada campo obligatorio según el tipo
    for campo in campos_requeridos:
        if campo == "tecnico":
            if not tecnico:
                campos_faltantes.append("Técnico Responsable")
        elif campo == "cotizacion":
            if not cotizacion:
                campos_faltantes.append("Cotización")
        elif campo == "orden_compra":
            if not orden_compra:
                campos_faltantes.append("Orden de Compra")
        elif campo == "factura":
            if not factura:
                campos_faltantes.append("Número de Factura")
        elif campo == "ppm":
            if not ppm:
                campos_faltantes.append("PPM Agua")
        elif campo == "corriente":
            if not corriente:
                campos_faltantes.append("Corriente de Trabajo")
        elif campo == "voltaje":
            if not voltaje:
                campos_faltantes.append("Voltaje de Trabajo")
        elif campo == "presion_entrada":
            if not presion_entrada:
                campos_faltantes.append("Presión de Entrada")
        elif campo == "presion_salida":
            if not presion_salida:
                campos_faltantes.append("Presión de Salida")
        elif campo == "observaciones":
            if not observaciones:
                campos_faltantes.append("Observaciones Generales")
        elif campo == "lista_piezas":
            if not st.session_state.lista_piezas:
                campos_faltantes.append("Piezas y Partes")
        elif campo == "boquillas_instaladas":
            if boquillas_instaladas == 0:
                campos_faltantes.append("Boquillas Instaladas")
        elif campo == "fotos_antes":
            if not fotos_antes:
                campos_faltantes.append("Foto ANTES de la mantención")
        elif campo == "fotos_despues":
            if not fotos_despues:
                campos_faltantes.append("Foto DESPUÉS de la mantención")
        elif campo == "capacitacion_si":
            if capacitacion_si == "No":
                campos_faltantes.append("Capacitación")
        elif campo == "recibe_nombre":
            if not recibe_nombre:
                campos_faltantes.append("Nombre Cliente")
        elif campo == "recibe_email":
            if not recibe_email:
                campos_faltantes.append("Email")
        elif campo == "recibe_cargo":
            if not recibe_cargo:
                campos_faltantes.append("Cargo")
        elif campo == "recibe_telefono":
            if not recibe_telefono:
                campos_faltantes.append("Teléfono")

    if campos_faltantes:
        st.warning(f"⚠️ Por favor, complete los siguientes campos obligatorios para {tipo}:\n" + "\n".join(
            [f"- {campo}" for campo in campos_faltantes]))
    else:
        # Aquí continúa el código existente para generar el Excel
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb["IT"]

        ws["S4"] = it_num
        ws["O7"] = cotizacion
        ws["P7"] = orden_compra
        ws["Q7"] = factura
        ws["R7"] = tipo
        ws["S7"] = fecha.strftime("%d/%m/%Y")
        ws["S10"] = serie
        ws["P12"] = ppm
        ws["S9"] = modelo
        ws["P9"] = razon_social
        ws["P10"] = rut_cliente
        ws["P11"] = ubicacion
        ws["P14"] = corriente
        ws["P13"] = voltaje
        ws["S13"] = presion_entrada
        ws["S14"] = presion_salida
        ws["O49"] = observaciones
        #ws["S37"] = fotos
        ws["P62"] = tecnico_upper
        ws["P58"] = recibe_nombre
        ws["S58"] = recibe_cargo
        ws["P59"] = recibe_email
        ws["P60"] = recibe_telefono
        ws["P15"] = str(boquillas_instaladas)
        ws["S11"]  = serierack
        ws["S12"] = agua

        for idx, (n, r) in enumerate(st.session_state.cap_list):
            if idx < 3:
                ws[f"P{51+idx}"] = n
                ws[f"S{51+idx}"] = r

        # === Actualización de las piezas en el Excel ===
        start_row = 18
        if st.session_state.lista_piezas:
            for i, pieza in enumerate(st.session_state.lista_piezas):
                row_idx = start_row + i
                if row_idx > 24:  # Límite de filas
                    break
                ws[f"O{row_idx}"] = pieza["codigo"]
                ws[f"P{row_idx}"] = pieza["nombre"]
                ws[f"R{row_idx}"] = str(pieza["cantidad"])
                ws[f"S{row_idx}"] = pieza["garantia"]


        if fotos_antes is not None:
            img = Image.open(fotos_antes)
            img.thumbnail((400, 400))
            bio = BytesIO()
            img.save(bio, format="PNG")
            img_excel = OpenpyxlImage(bio)
            ws.add_image(img_excel, "O37")

        if fotos_despues is not None:
            img = Image.open(fotos_despues)
            img.thumbnail((400, 400))
            bio = BytesIO()
            img.save(bio, format="PNG")
            img_excel = OpenpyxlImage(bio)
            ws.add_image(img_excel, "Q37")

        # Cuando proceses la firma para guardarla en Excel:
        if sig_canvas.image_data is not None:
            # Convertir los datos de la imagen a un array numpy
            img_data = sig_canvas.image_data

            # Crear una imagen con mayor resolución
            img = Image.fromarray(img_data.astype('uint8'), 'RGBA')

            # Aumentar el tamaño de la imagen para mejor calidad
            new_width = int(img.width * 0.9)      #alto
            new_height = int(img.height * 0.9)    #ancho
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)

            # Mejorar el contraste de la firma
            from PIL import ImageEnhance

            # Aumentar el contraste
            enhancer = ImageEnhance.Contrast(img)
            img = enhancer.enhance(1.5)  # Ajusta el valor entre 1.0 y 2.0 para más o menos contraste

            # Aumentar la nitidez
            enhancer = ImageEnhance.Sharpness(img)
            img = enhancer.enhance(1.3)  # Ajusta el valor entre 1.0 y 2.0 para más o menos nitidez

            # Guardar la imagen con alta calidad
            img_byte_arr = BytesIO()
            img.save(img_byte_arr, format='PNG', quality=95, optimize=True)
            img_byte_arr = img_byte_arr.getvalue()

            # Cuando agregas la imagen al Excel:
            img_excel = OpenpyxlImage(BytesIO(img_byte_arr))
            ws.add_image(img_excel, "P61")

        output = BytesIO()
        wb.active = wb["IT"]
        wb.save(output)
        output.seek(0)

        nombre_archivo = f"Informe_Tecnico_{serie or 'sin_serie'}.xlsx"
        st.success("✅ Informe generado exitosamente")
        st.download_button(
            label="📤 Descargar Informe",
            data=output,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
